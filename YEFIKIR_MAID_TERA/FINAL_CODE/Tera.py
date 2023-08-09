def tera(file, calendar, members):
    # importing the necessary libraries
    from openpyxl import load_workbook
    import xlsxwriter
    import ast
    import json
    from ethiopian_date import EthiopianDateConverter
    # import calendar
    from datetime import datetime
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Font

    # reading file1 sheet1
    workbook1 = load_workbook(filename=file)
    sheet1 = workbook1[calendar]
    sheet2 = workbook1[members]
    clmninfo1 = {}
    clmninfo2 = {}
    wendoch, setoch, beal, kenbewchi, kutr = '', '', '', '', ''
    year = 2015

    new = 'ሰኔ ሐምሌ ነሐሴ'.split()

    for indx, column1 in enumerate(sheet1.iter_cols()):

        # example: clmninfo1['Asset Name'] = ['J0', 'J', 'Asset Name']
        clmninfo1[column1[0].value] = [column1[0].coordinate,
                                       column1[0].column_letter, column1[0].value, indx+1]

    for column1 in (sheet2.iter_cols()):

        # example: clmninfo1['Asset Name'] = ['J0', 'J', 'Asset Name']
        clmninfo2[column1[0].value] = [column1[0].coordinate,
                                       column1[0].column_letter, column1[0].value]

    for wer in new:
        for indx, value in enumerate(sheet1[clmninfo1[wer][1]][1::]):
            wendoch = 'ወንድ'
            setoch = 'ሴት'
            if value.value != None:
                kutr = value.value.split('|')[5].split()[0]
                ken = str(EthiopianDateConverter.to_gregorian(
                    year, clmninfo1[wer][-1], int(kutr))).split('-')
                kenbewchi = datetime(int(ken[0]), int(ken[1]), int(ken[2]))
                kenbewchi = f"{kenbewchi.strftime('%B')} {ken[2]}"
                beal = beal = value.value.split('|')[5]
                zkr = f"""ዘካሪ፥ {wendoch} እና {setoch}\nየሚዘከረው በዓል እና ቀኑ፥  {beal}\nጸበል ጻዲቅ የሚቀርብበት ቀን በውጭ:  {kenbewchi}\nጸበል ጻዲቅ የሚቀርብበት ቀን በኢትዮጵያ፥ {wer}  {kutr}"""
                loc = f"{clmninfo1[wer][1]}{indx+2}"
                sheet1[loc].value = zkr
                sheet1[loc].font = Font(bold=True)
                sheet1[loc].fill = PatternFill(
                    patternType='solid', fgColor='9BC2E6')

    Resultfilenamee = '{}{}.xlsx'.format(
        file[0:len(file)-5], '_____Modified_Result')

    # saving the file as 'newfilenamee'
    workbook1.save(filename=Resultfilenamee)
    workbook1.close()

    return  # clmninfo1, clmninfo2


file = '/Users/ab/Desktop/JA-COLAB/JA-COLAB/YEFIKIR_MAID_TERA/UNDER_DEVELOPMENT/የየእሑድ_ዝክር_የበዓላት_ዝርዝር_የዓመት.xlsx'
calendar = 'የየእሑድ ዝክር የበዓላት ዝርዝር (2)'
members = 'አባላት'


if __name__ == "__main__":
    tera(file, calendar, members)



# ተራ ለቤቲ በንቲ እና ስንታየሁ አንድላይ መዘከር ይፈልጋሉ (ሁለት ቀን ቢሆን ተመራጭ ነው)
# add  a feature to swap two assigned members
# add  a feature to swap reserve to assigned members
# add a relational data that shows a percentile of how close a group of people are
# member should have: name in amharic, name  in english, phone, address
#                   , email, close friends, group, beal zikir list(amet,wer,ken), activity status (in percentage)
#                   , unpaid membership fee
